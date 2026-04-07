import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import subprocess

# Yeni Excel dosyasının yolunu belirleyin
file_name = 'exact_replication1.xlsx'
file_path = os.path.join(os.getcwd(), file_name)

# Yeni bir çalışma kitabı oluşturun ve aktif sayfayı seçin
workbook = openpyxl.Workbook()
sheet = workbook.active

# Hücre stillerini tanımlayın
header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
subheader_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Başlıklar ve birleştirilmiş hücreler
headers = [
    ("A1", "MALZEME TANIMI", "A1:F1"), 
    ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
    ("A3", "", ""),("B3", "", ""),("C3", "", ""),("D3", "", ""),("E3", "", ""),("F3", "", ""),
    ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
    ("A6", "1. Tünel Kemer", ""), ("A7", "1. Tünel İnvert", ""),("B6", "", ""),("B7", "", ""),
    ("A11", "", ""),("B11", "", ""),("C11", "", ""),("D11", "", ""),("E11", "", ""),("F11", "", ""),("G11", "", ""),("H11", "", ""),
    ("I11", "", ""),("J11", "", ""),("K11", "", ""),
    ("A16", "", ""),("B16", "", ""),("C16", "", ""),("D16", "", ""),
    ("A9", "1. TÜNEL ZEMİN YÜKLERİ", "A9:K9"), 
    ("A10", "Φ'", ""), ("B10", "m", ""), ("C10", "b", ""), ("D10", "B", ""), ("E10", "f", ""), ("F10", "h", ""), ("G10", "Y", ""), ("H10", "Y'", ""), ("I10", "K0", ""), ("J10", "EV", ""), ("K10", "EH", ""),
    ("A20", "", ""),("B20", "", ""),("C20", "", ""),("D20", "", ""),("E20", "", ""),("F20", "", ""),("G20", "", ""),("H20", "", ""),
    ("A13", "Yatak Kaysayıları", "A13:D13"), 
    ("C14", "1. Tünel","C14:D14"),
    ("A15", "E' (kPa)", ""), ("B15", "v", ""), ("C15", "R", ""), ("D15", "k", ""),
    ("A18", "1. TÜNEL DEPREM YÜKLERİ", "A18:H18"),
    ("A19", "S", ""), ("B19", "C", ""), ("C19", "Y", ""), ("D19", "H", ""), ("E19", "PGA (72)", ""), ("F19", "PGA (2475)", ""), ("G19", "S1", ""), ("H19", "S2", "")
]

# Başlıkları ve birleştirilmiş hücreleri uygula
for cell, value, merge in headers:
    sheet[cell] = value
    if "MALZEME TANIMI" in value or "KESİT KALINLIKLARI (cm)" in value or "TÜNEL ZEMİN YÜKLERİ" in value or "Yatak Kaysayıları" in value or "TÜNEL DEPREM YÜKLERİ" in value:
        sheet[cell].fill = header_fill
    else:
        sheet[cell].fill = subheader_fill
    sheet[cell].font = bold_font
    sheet[cell].alignment = center_alignment
    sheet[cell].border = thin_border
    if merge:
        sheet.merge_cells(merge)

# Sütun genişliklerini ayarla
column_widths = {'A': 20, 'B': 20, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12}
for col, width in column_widths.items():
    sheet.column_dimensions[col].width = width

# Belirli hücreleri sarı renkle doldur
yellow_cells = ["J11", "K11", "G20", "H20","D16"]
for cell in yellow_cells:
    sheet[cell].fill = yellow_fill

# Çalışma kitabını kaydet
workbook.save(file_path)

# Excel dosyasını aç
if os.name == 'posix':  # macOS veya Linux
    subprocess.run(['open', file_path])
elif os.name == 'nt':  # Windows
    os.startfile(file_path)

print(f"Biçimlendirilmiş Excel sayfası {file_path} konumunda oluşturuldu.")
