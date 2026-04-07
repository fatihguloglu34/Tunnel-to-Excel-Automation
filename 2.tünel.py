import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import subprocess

# Yeni Excel dosyasının yolunu belirleyin
file_name = 'exact_replication2.xlsx'
file_path = os.path.join(os.getcwd(), file_name)

# Yeni bir çalışma kitabı oluşturun ve aktif sayfayı seçin
workbook = openpyxl.Workbook()
sheet = workbook.active

# Hücre stillerini tanımlayın
header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
subheader_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Başlıklar ve birleştirilmiş hücreler
headers = [
    ("A1", "MALZEME TANIMI", "A1:F1"), 
    ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
    ("A3", "", ""), ("B3", "", ""), ("C3", "", ""), ("D3", "", ""), ("E3", "", ""), ("F3", "", ""),
    ("B6", "", ""), ("B7", "", ""), ("B8", "", ""),("B9", "", ""),   
    ("A13", "", ""), ("B13", "", ""),("C13", "", ""),("D13", "", ""),("E13", "", ""),("F13", "", ""),("G13", "", ""), ("H13", "", ""),("I13", "", ""),("J13", "", ""),("K13", "", ""),
    ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
    ("A6", "1. Tünel Kemer", ""), ("A7", "1. Tünel İnvert", ""), ("A8", "2. Tünel Kemer", ""), ("A9", "2. Tünel İnvert", ""),
    ("A17", "", ""), ("B17", "", ""), ("C17", "", ""), ("D17", "", ""), ("E17", "", ""), ("F17", "", ""), ("G17", "", ""), ("H17", "", ""), ("I17", "", ""), ("J17", "", ""), ("K17", "", ""),
    ("E21", "R", ""),("F21", "k", ""),("E23", "", ""),
    ("A11", "1. TÜNEL ZEMİN YÜKLERİ", "A11:K11"), 
    ("A12", "Φ'", ""), ("B12", "m", ""), ("C12", "b", ""), ("D12", "B", ""), ("E12", "f", ""), ("F12", "h", ""), ("G12", "Y", ""), ("H12", "Y'", ""), ("I12", "K0", ""), ("J12", "EV", ""), ("K12", "EH", ""),
    ("A22", "", ""),("B22", "", ""),("C22", "", ""),("D22", "", ""),("F22", "", ""),    
    ("A15", "2. TÜNEL ZEMİN YÜKLERİ", "A15:K15"), 
    ("A16", "Φ'", ""), ("B16", "m", ""), ("C16", "b", ""), ("D16", "B", ""), ("E16", "f", ""), ("F16", "h", ""), ("G16", "Y", ""), ("H16", "Y'", ""), ("I16", "K0", ""), ("J16", "EV", ""), ("K16", "EH", ""),
    ("A26", "", ""), ("B26", "", ""), ("C26", "", ""), ("D26", "", ""), ("E26", "", ""), ("F26", "", ""), ("G26", "", ""), ("H26", "", ""),
    ("A30", "", ""), ("B30", "", ""), ("C30", "", ""), ("D30", "", ""), ("E30", "", ""), ("F30", "", ""), ("H30", "", ""), ("G30", "", ""), 
    ("A19", "Yatak Kaysayıları", "A19:f19"), 
    ("A21", "E' (kPa)", ""), ("B21", "v", ""), ("C20", "1. Tünel", "C20:D20"),("E20", "2. Tünel", "E20:F20"),
    ("C21", "R", ""),("D21", "k", ""),
    ("A24", "1. TÜNEL DEPREM YÜKLERİ", "A24:H24"),
    ("A25", "S", ""), ("B25", "C", ""), ("C25", "Y", ""), ("D25", "H", ""), ("E25", "PGA (72)", ""), ("F25", "PGA (2475)", ""), ("G25", "S1", ""), ("H25", "S2", ""),
    ("A28", "2. TÜNEL DEPREM YÜKLERİ", "A28:H28"),
    ("A29", "S", ""), ("B29", "C", ""), ("C29", "Y", ""), ("D29", "H", ""), ("E29", "PGA (72)", ""), ("F29", "PGA (2475)", ""), ("G29", "S1", ""), ("H29", "S2", "")
]

# Başlıkları ve birleştirilmiş hücreleri uygula
for cell, value, merge in headers:
    sheet[cell] = value
    if "MALZEME TANIMI" in value or "KESİT KALINLIKLARI (cm)" in value or "TÜNEL ZEMİN YÜKLERİ" in value or "Yatak Kaysayıları" in value or "TÜNEL DEPREM YÜKLERİ" in value:
        sheet[cell].fill = header_fill
    else:
        sheet[cell].fill = subheader_fill
    sheet[cell].font = bold_font
    sheet[cell].alignment = center_alignment
    sheet[cell].border = thin_border
    if merge:
        sheet.merge_cells(merge)

# Sütun genişliklerini ayarla
column_widths = {'A': 20, 'B': 20, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12}
for col, width in column_widths.items():
    sheet.column_dimensions[col].width = width

# Belirli hücreleri sarı renkle doldur
yellow_cells = ["J13", "K13", "J17", "K17","G30","H30","G26","H26","D22","F22"]
for cell in yellow_cells:
    sheet[cell].fill = yellow_fill

# Çalışma kitabını kaydet
workbook.save(file_path)

# Excel dosyasını aç
if os.name == 'posix':  # macOS veya Linux
    subprocess.run(['open', file_path])
elif os.name == 'nt':  # Windows
    os.startfile(file_path)

print(f"Biçimlendirilmiş Excel sayfası {file_path} konumunda oluşturuldu.")
