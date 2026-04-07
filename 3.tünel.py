import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import subprocess

# Define the path for the new Excel file
file_name = 'exact_replication_3.xlsx'
file_path = os.path.join(os.getcwd(), file_name)

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define cell styles
header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
subheader_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Headers and merged cells
headers = [
    ("A1", "MALZEME TANIMI", "A1:F1"), 
    ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
    ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
    ("A3", "", ""),("B3", "", ""),("C3", "", ""),("D3", "", ""),("E3", "", ""),("F3", "", ""),
    ("A6", "1. Tünel Kemer", ""), ("B6", "", ""),  # Boş sütun
    ("A7", "1. Tünel İnvert", ""), ("B7", "", ""),  # Boş sütun
    ("A8", "2. Tünel Kemer", ""), ("B8", "", ""),  # Boş sütun
    ("A9", "2. Tünel İnvert", ""), ("B9", "", ""),  # Boş sütun
    ("A10", "3. Tünel Kemer", ""), ("B10", "", ""),  # Boş sütun
    ("A11", "3. Tünel İnvert", ""), ("B11", "", ""),  # Boş sütun
    ("A12", "", ""), ("B12", "", ""),  # Boş satır
    ("A13", "1. TÜNEL ZEMİN YÜKLERİ", "A13:K13"), 
    ("A14", "Φ'", ""), ("B14", "m", ""), ("C14", "b", ""), ("D14", "B", ""), ("E14", "f", ""), ("F14", "h", ""), ("G14", "Y", ""), ("H14", "Y'", ""), ("I14", "K0", ""), ("J14", "EV", ""), ("K14", "EH", ""),
    ("A15", "", ""), ("B15", "", ""), ("C15", "", ""), ("D15", "", ""), ("E15", "", ""), ("F15", "", ""), ("G15", "", ""), ("H15", "", ""), ("I15", "", ""), ("J15", "", ""), ("K15", "", ""),
    
    ("A17", "2. TÜNEL ZEMİN YÜKLERİ", "A17:K17"), 
    ("A18", "Φ'", ""), ("B18", "m", ""), ("C18", "b", ""), ("D18", "B", ""), ("E18", "f", ""), ("F18", "h", ""), ("G18", "Y", ""), ("H18", "Y'", ""), ("I18", "K0", ""), ("J18", "EV", ""), ("K18", "EH", ""),
    ("A19", "", ""), ("B19", "", ""), ("C19", "", ""), ("D19", "", ""), ("E19", "", ""), ("F19", "", ""), ("G19", "", ""), ("H19", "", ""), ("I19", "", ""), ("J19", "", ""), ("K19", "", ""),

    ("A21", "3. TÜNEL ZEMİN YÜKLERİ", "A21:K21"), 
    ("A22", "Φ'", ""), ("B22", "m", ""), ("C22", "b", ""), ("D22", "B", ""), ("E22", "f", ""), ("F22", "h", ""), ("G22", "Y", ""), ("H22", "Y'", ""), ("I22", "K0", ""), ("J22", "EV", ""), ("K22", "EH", ""),
    ("A23", "", ""), ("B23", "", ""), ("C23", "", ""), ("D23", "", ""), ("E23", "", ""), ("F23", "", ""), ("G23", "", ""), ("H23", "", ""), ("I23", "", ""), ("J23", "", ""), ("K23", "", ""),
    
    ("A25", "Yatak Kaysayıları", "A25:H25"),
    ("A27", "E' (kPa)", ""), ("B27", "v", ""), 
     ("C27", "R", ""), ("D27", "k", ""), ("C28", "", ""), 
    ("E27", "R", ""), ("F27", "k", ""), ("E28", "", ""),
     ("G27", "R", ""), ("H27", "k", ""), ("G28", "", ""), ("G26", "3. Tünel", "G26:H26"),
     ("E26", "2. Tünel", "E26:D26"),
     ("C26", "1. Tünel", "C26:D26"),
    ("A28", "", ""), ("B28", "", ""), ("C28", "", ""), ("D28", "", ""), ("E28", "", ""), ("F28", "", ""), ("G28", "", ""), ("H28", "", ""),
    ("A30", "1. TÜNEL DEPREM YÜKLERİ", "A30:H30"),
    ("A31", "S", ""), ("B31", "C", ""), ("C31", "Y", ""), ("D31", "H", ""), ("E31", "PGA (72)", ""), ("F31", "PGA (2475)", ""), ("G31", "S1", ""), ("H31", "S2", ""),
    ("A32", "", ""), ("B32", "", ""), ("C32", "", ""), ("D32", "", ""), ("E32", "", ""), ("F32", "", ""), ("G32", "", ""), ("H32", "", ""),

    ("A34", "2. TÜNEL DEPREM YÜKLERİ", "A34:H34"),
    ("A35", "S", ""), ("B35", "C", ""), ("C35", "Y", ""), ("D35", "H", ""), ("E35", "PGA (72)", ""), ("F35", "PGA (2475)", ""), ("G35", "S1", ""), ("H35", "S2", ""),
    ("A36", "", ""), ("B36", "", ""), ("C36", "", ""), ("D36", "", ""), ("E36", "", ""), ("F36", "", ""), ("G36", "", ""), ("H36", "", ""),

    ("A38", "3. TÜNEL DEPREM YÜKLERİ", "A38:H38"),
    ("A39", "S", ""), ("B39", "C", ""), ("C39", "Y", ""), ("D39", "H", ""), ("E39", "PGA (72)", ""), ("F39", "PGA (2475)", ""), ("G39", "S1", ""), ("H39", "S2", ""),
    ("A40", "", ""), ("B40", "", ""), ("C40", "", ""), ("D40", "", ""), ("E40", "", ""), ("F40", "", ""), ("G40", "", ""), ("H40", "", ""),
]

# Apply headers
for cell, value, merge in headers:
    sheet[cell] = value
    if "MALZEME TANIMI" in value or "KESİT KALINLIKLARI (cm)" in value or "ZEMİN YÜKLERİ" in value or "Yatak Kaysayıları" in value or "TÜNEL DEPREM YÜKLERİ" in value:
        sheet[cell].fill = header_fill
    else:
        sheet[cell].fill = subheader_fill
    sheet[cell].font = bold_font
    sheet[cell].alignment = center_alignment
    sheet[cell].border = thin_border
    if merge:
        sheet.merge_cells(merge)

# Set column widths
column_widths = {'A': 20, 'B': 20, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12}
for col, width in column_widths.items():
    sheet.column_dimensions[col].width = width

# Fill specific cells with yellow color
yellow_cells = ["K15", "J19", "K23", "J15", "J23","H40","H36", "G36","K19","G40","D28","F28","H28","G32","H32"]
for cell in yellow_cells:
    sheet[cell].fill = yellow_fill

# Save the workbook
workbook.save(file_path)

print(f"The formatted Excel sheet has been created at {file_path}.")

# Open the Excel file
if os.name == 'posix':  # macOS or Linux
    subprocess.run(['open', file_path])
elif os.name == 'nt':  # Windows
    os.startfile(file_path)
