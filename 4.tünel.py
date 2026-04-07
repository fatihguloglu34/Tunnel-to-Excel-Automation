import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import subprocess

# Define the path for the new Excel file
file_name = 'exact_replication4.xlsx'
file_path = os.path.join(os.getcwd(), file_name)

# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Define cell styles
header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
subheader_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
bold_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Headers and merged cells
headers = [
    ("A1", "MALZEME TANIMI", "A1:F1"), 
    ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
    ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
    ("A6", "1. Tünel Kemer", ""), ("B6", "", ""),  # Boş sütun
    ("A7", "1. Tünel İnvert", ""), ("B7", "", ""),  # Boş sütun
    ("A8", "2. Tünel Kemer", ""), ("B8", "", ""),  # Boş sütun
    ("A9", "2. Tünel İnvert", ""), ("B9", "", ""),  # Boş sütun
    ("A10", "3. Tünel Kemer", ""), ("B10", "", ""),  # Boş sütun
    ("A11", "3. Tünel İnvert", ""), ("B11", "", ""),  # Boş sütun
    ("A12", "4. Tünel Kemer", ""), ("B12", "", ""),  # Boş sütun
    ("A13", "4. Tünel İnvert", ""), ("B13", "", ""),  # Boş sütun
    ("A15", "1. TÜNEL ZEMİN YÜKLERİ", "A15:K15"), 
    ("A16", "Φ'", ""), ("B16", "m", ""), ("C16", "b", ""), ("D16", "B", ""), ("E16", "f", ""), ("F16", "h", ""), ("G16", "Y", ""), ("H16", "Y'", ""), ("I16", "K0", ""), ("J16", "EV", ""), ("K16", "EH", ""),
    ("A17", "", ""), ("B17", "", ""), ("C17", "", ""), ("D17", "", ""), ("E17", "", ""), ("F17", "", ""), ("G17", "", ""), ("H17", "", ""), ("I17", "", ""), ("J17", "", ""), ("K17", "", ""),
    ("A19", "2. TÜNEL ZEMİN YÜKLERİ", "A19:K19"), 
    ("A20", "Φ'", ""), ("B20", "m", ""), ("C20", "b", ""), ("D20", "B", ""), ("E20", "f", ""), ("F20", "h", ""), ("G20", "Y", ""), ("H20", "Y'", ""), ("I20", "K0", ""), ("J20", "EV", ""), ("K20", "EH", ""),
    ("A21", "", ""), ("B21", "", ""), ("C21", "", ""), ("D21", "", ""), ("E21", "", ""), ("F21", "", ""), ("G21", "", ""), ("H21", "", ""), ("I21", "", ""), ("J21", "", ""), ("K21", "", ""),
    ("A23", "3. TÜNEL ZEMİN YÜKLERİ", "A23:K23"), 
    ("A24", "Φ'", ""), ("B24", "m", ""), ("C24", "b", ""), ("D24", "B", ""), ("E24", "f", ""), ("F24", "h", ""), ("G24", "Y", ""), ("H24", "Y'", ""), ("I24", "K0", ""), ("J24", "EV", ""), ("K24", "EH", ""),
    ("A25", "", ""), ("B25", "", ""), ("C25", "", ""), ("D25", "", ""), ("E25", "", ""), ("F25", "", ""), ("G25", "", ""), ("H25", "", ""), ("I25", "", ""), ("J25", "", ""), ("K25", "", ""),
    ("A27", "4. TÜNEL ZEMİN YÜKLERİ", "A27:K27"), 
    ("A28", "Φ'", ""), ("B28", "m", ""), ("C28", "b", ""), ("D28", "B", ""), ("E28", "f", ""), ("F28", "h", ""), ("G28", "Y", ""), ("H28", "Y'", ""), ("I28", "K0", ""), ("J28", "EV", ""), ("K28", "EH", ""),
    ("A29", "", ""), ("B29", "", ""), ("C29", "", ""), ("D29", "", ""), ("E29", "", ""), ("F29", "", ""), ("G29", "", ""), ("H29", "", ""), ("I29", "", ""), ("J29", "", ""), ("K29", "", ""),
    ("A31", "Yatak Kaysayıları", "A31:J31"),
    ("A33", "E' (kPa)", ""), ("B33", "v", ""), 
    ("C32", "1. Tünel", "C32:D32"), ("C33", "R", ""), ("D33", "k", ""), ("C34", "", ""), 
    ("E32", "2. Tünel", "E32:F32"), ("E33", "R", ""), ("F33", "k", ""), ("E34", "", ""),
    ("G32", "3. Tünel", "G32:H32"), ("G33", "R", ""), ("H33", "k", ""), ("G34", "", ""),
    ("I32", "4. Tünel", ""), ("I33", "R", ""), ("J33", "k", ""), ("I34", "", ""), ("J32", "", ""), ("J34", "", ""),
    ("A3", "", ""), ("B3", "", ""), ("C3", "", ""), ("D3", "", ""),("E3", "", ""),  ("F3", "", ""), 
    ("A36", "1. TÜNEL DEPREM YÜKLERİ", "A36:H36"),
    ("A37", "S", ""), ("B37", "C", ""), ("C37", "Y", ""), ("D37", "H", ""), ("E37", "PGA (72)", ""), ("F37", "PGA (2475)", ""), ("G37", "S1", ""), ("H37", "S2", ""),
    ("A38", "", ""), ("B38", "", ""), ("C38", "", ""), ("D38", "", ""), ("E38", "", ""), ("F38", "", ""), ("G38", "", ""), ("H38", "", ""),
    ("A40", "2. TÜNEL DEPREM YÜKLERİ", "A40:H40"),
    ("A34", "", ""),("B34", "", ""),("D34", "", ""),("F34", "", ""),("H34", "", ""),("A38", "", ""),
    ("A41", "S", ""), ("B41", "C", ""), ("C41", "Y", ""), ("D41", "H", ""), ("E41", "PGA (72)", ""), ("F41", "PGA (2475)", ""), ("G41", "S1", ""), ("H41", "S2", ""),
    ("A42", "", ""), ("B42", "", ""), ("C42", "", ""), ("D42", "", ""), ("E42", "", ""), ("F42", "", ""), ("G42", "", ""), ("H42", "", ""),
    ("A44", "3. TÜNEL DEPREM YÜKLERİ", "A44:H44"),
    ("A45", "S", ""), ("B45", "C", ""), ("C45", "Y", ""), ("D45", "H", ""), ("E45", "PGA (72)", ""), ("F45", "PGA (2475)", ""), ("G45", "S1", ""), ("H45", "S2", ""),
    ("A46", "", ""), ("B46", "", ""), ("C46", "", ""), ("D46", "", ""), ("E46", "", ""), ("F46", "", ""), ("G46", "", ""), ("H46", "", ""),
    ("A48", "4. TÜNEL DEPREM YÜKLERİ", "A48:H48"),
    ("A49", "S", ""), ("B49", "C", ""), ("C49", "Y", ""), ("D49", "H", ""), ("E49", "PGA (72)", ""), ("F49", "PGA (2475)", ""), ("G49", "S1", ""), ("H49", "S2", ""),
    ("A50", "", ""), ("B50", "", ""), ("C50", "", ""), ("D50", "", ""), ("E50", "", ""), ("F50", "", ""), ("G50", "", ""), ("H50", "", ""),
]



# Apply headers
for cell, value, merge in headers:
    sheet[cell] = value
    if "MALZEME TANIMI" in value or "KESİT KALINLIKLARI (cm)" in value or "ZEMİN YÜKLERİ" in value or "Yatak Kaysayıları" in value or "TÜNEL DEPREM YÜKLERİ" in value:
        sheet[cell].fill = header_fill
    else:
        sheet[cell].fill = subheader_fill
    sheet[cell].font = bold_font
    sheet[cell].alignment = center_alignment
    sheet[cell].border = thin_border
    if merge:
        sheet.merge_cells(merge)

# Set column widths
column_widths = {'A': 20, 'B': 20, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12}
for col, width in column_widths.items():
    sheet.column_dimensions[col].width = width

# Fill specific cells with yellow color
yellow_cells = ["K17", "J21", "K25", "J17", "J25","H40", "D34","F34","H34","J34","H36", "G36", "H48", "G48","K21","G44","H44","G40","J29","K29","G38","H38","G42","H42","G46","H46","G50","H50"]
for cell in yellow_cells:
    sheet[cell].fill = yellow_fill

# Save the workbook
workbook.save(file_path)

print(f"The formatted Excel sheet has been created at {file_path}.")

# Open the Excel file
if os.name == 'posix':  # macOS or Linux
    subprocess.run(['open', file_path])
elif os.name == 'nt':  # Windows
    os.startfile(file_path)
