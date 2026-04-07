import sys
import os
import subprocess
from PyQt5 import QtWidgets, uic
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import range_boundaries

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        ui_path = os.path.expanduser("~/Desktop/kesit3.ui")
        uic.loadUi(ui_path, self)

        # Qt Designer'da tanımlanan widget'ları bulma ve bağlama
        self.comboBox_beton = self.findChild(QtWidgets.QComboBox, 'Betonsinifbuton')
        self.comboBox_yonetmelik = self.findChild(QtWidgets.QComboBox, 'yonetmelikbuton')
        self.pushButton = self.findChild(QtWidgets.QPushButton, 'pushButton')
        self.lineEdit_tunnel = self.findChild(QtWidgets.QLineEdit, 'tunelsayisibuton')

        # Giriş alanlarını ve etiketlerini saklayacak liste
        self.tunnel_inputs = [
            (self.findChild(QtWidgets.QLabel, 'birincitunelkemer'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton')),
            (self.findChild(QtWidgets.QLabel, 'birincitunelinvert'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton_2')),
            (self.findChild(QtWidgets.QLabel, 'ikincitunelkemer'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton_3')),
            (self.findChild(QtWidgets.QLabel, 'ikincitunelinvert'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton_4')),
            (self.findChild(QtWidgets.QLabel, 'ucuncutunelkemer'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton_5')),
            (self.findChild(QtWidgets.QLabel, 'ucuncutunelinvert'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton_6')),
            (self.findChild(QtWidgets.QLabel, 'dorduncutunelkemer'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton_10')),
            (self.findChild(QtWidgets.QLabel, 'dorduncutunelkemerinvert'), self.findChild(QtWidgets.QLineEdit, 'birincitunelkemerbuton_8'))
        ]

        # Tünel sayısı değişikliklerini dinlemek için bağlantı
        self.lineEdit_tunnel.textChanged.connect(self.update_tunnel_inputs)

        # Buton bağlantısını ayarlama
        self.pushButton.clicked.connect(self.write_to_excel)

        # İlk yüklemede tünel girişlerini güncelle
        self.update_tunnel_inputs()

    def update_tunnel_inputs(self):
        try:
            tunnel_count = int(self.lineEdit_tunnel.text())
        except ValueError:
            tunnel_count = 0

        for i, (label, input_field) in enumerate(self.tunnel_inputs):
            if i < tunnel_count * 2:
                label.setVisible(True)
                input_field.setVisible(True)
            else:
                label.setVisible(False)
                input_field.setVisible(False)

    def get_elasticity_modulus(self, beton_sinifi, yonetmelik):
        if yonetmelik == "TS500":
            if beton_sinifi == "C25":
                return 30000000
            elif beton_sinifi == "C30":
                return 32000000
            elif beton_sinifi == "C35":
                return 33000000
            elif beton_sinifi == "C40":
                return 34000000
            elif beton_sinifi == "C45":
                return 36000000
            elif beton_sinifi == "C50":
                return 37000000
        elif yonetmelik == "EUROCODE":
            if beton_sinifi == "C25":
                return 31000000
            elif beton_sinifi == "C30":
                return 33000000
            elif beton_sinifi == "C35":
                return 35000000
            elif beton_sinifi == "C40":
                return 35000000
            elif beton_sinifi == "C45":
                return 36000000
            elif beton_sinifi == "C50":
                return 37000000
        return 0

    def get_start_cell(self, sheet, cell):
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
                return sheet.cell(row=min_row, column=min_col)
        return cell

    def write_to_excel(self):
        try:
            tünel_sayisi = int(self.lineEdit_tunnel.text())
        except ValueError:
            tünel_sayisi = 0

        if tünel_sayisi == 1:
            file_name = 'exact_replication1.xlsx'
            headers = [
                   ("A1", "MALZEME TANIMI", "A1:F1"), 
      ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
      ("A3", "", ""),("B3", "", ""),("C3", "", ""),("D3", "", ""),("E3", "", ""),("F3", "", ""),
      ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
      ("A6", "1. Tünel Kemer", ""), ("A7", "1. Tünel İnvert", ""),("B6", "", ""),("B7", "", ""),
      ("A11", "", ""),("B11", "", ""),("C11", "", ""),("D11", "", ""),("E11", "", ""),("F11", "", ""),("G11", "", ""),("H11", "", ""),
      ("I11", "", ""),("J11", "", ""),("K11", "", ""),
      ("A16", "", ""),("B16", "", ""),("C16", "", ""),("D16", "", ""),
      ("A9", "1. TÜNEL ZEMİN YÜKLERİ", "A9:K9"), 
      ("A10", "Φ'", ""), ("B10", "m", ""), ("C10", "b", ""), ("D10", "B", ""), ("E10", "f", ""), ("F10", "h", ""), ("G10", "Y", ""), ("H10", "Y'", ""), ("I10", "K0", ""), ("J10", "EV", ""), ("K10", "EH", ""),
      ("A20", "", ""),("B20", "", ""),("C20", "", ""),("D20", "", ""),("E20", "", ""),("F20", "", ""),("G20", "", ""),("H20", "", ""),
      ("A13", "Yatak Kaysayıları", "A13:D13"), 
      ("C14", "1. Tünel", "C14:D14"),
      ("A15", "E' (kPa)", ""), ("B15", "v", ""), ("C15", "R", ""), ("D15", "k", ""),
      ("A18", "1. TÜNEL DEPREM YÜKLERİ", "A18:H18"),
      ("A19", "S", ""), ("B19", "C", ""), ("C19", "Y", ""), ("D19", "H", ""), ("E19", "PGA (72)", ""), ("F19", "PGA (2475)", ""), ("G19", "S1", ""), ("H19", "S2", "")
            ]
            yellow_cells = ["J11", "K11", "G20", "H20","D16"]

        elif tünel_sayisi == 2:
            file_name = 'exact_replication2.xlsx'
            headers = [
                ("A1", "MALZEME TANIMI", "A1:F1"), 
    ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
    ("A3", "", ""), ("B3", "", ""), ("C3", "", ""), ("D3", "", ""), ("E3", "", ""), ("F3", "", ""),
    ("B6", "", ""), ("B7", "", ""), ("B8", "", ""),("B9", "", ""),   
    ("A13", "", ""), ("B13", "", ""),("C13", "", ""),("D13", "", ""),("E13", "", ""),("F13", "", ""),("G13", "", ""), ("H13", "", ""),("I13", "", ""),("J13", "", ""),("K13", "", ""),
    ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
    ("A6", "1. Tünel Kemer", ""), ("A7", "1. Tünel İnvert", ""), ("A8", "2. Tünel Kemer", ""), ("A9", "2. Tünel İnvert", ""),
    ("A17", "", ""), ("B17", "", ""), ("C17", "", ""), ("D17", "", ""), ("E17", "", ""), ("F17", "", ""), ("G17", "", ""), ("H17", "", ""), ("I17", "", ""), ("J17", "", ""), ("K17", "", ""),
    ("E21", "R", ""),("F21", "k", ""),("E23", "", ""),
    ("A11", "1. TÜNEL ZEMİN YÜKLERİ", "A11:K11"), 
    ("A12", "Φ'", ""), ("B12", "m", ""), ("C12", "b", ""), ("D12", "B", ""), ("E12", "f", ""), ("F12", "h", ""), ("G12", "Y", ""), ("H12", "Y'", ""), ("I12", "K0", ""), ("J12", "EV", ""), ("K12", "EH", ""),
    ("A22", "", ""),("B22", "", ""),("C22", "", ""),("D22", "", ""),("F22", "", ""),    
    ("A15", "2. TÜNEL ZEMİN YÜKLERİ", "A15:K15"), 
    ("A16", "Φ'", ""), ("B16", "m", ""), ("C16", "b", ""), ("D16", "B", ""), ("E16", "f", ""), ("F16", "h", ""), ("G16", "Y", ""), ("H16", "Y'", ""), ("I16", "K0", ""), ("J16", "EV", ""), ("K16", "EH", ""),
    ("A26", "", ""), ("B26", "", ""), ("C26", "", ""), ("D26", "", ""), ("E26", "", ""), ("F26", "", ""), ("G26", "", ""), ("H26", "", ""),
    ("A30", "", ""), ("B30", "", ""), ("C30", "", ""), ("D30", "", ""), ("E30", "", ""), ("F30", "", ""), ("H30", "", ""), ("G30", "", ""), 
    ("A19", "Yatak Kaysayıları", "A19:f19"), 
    ("A21", "E' (kPa)", ""), ("B21", "v", ""), ("C20", "1. Tünel", "C20:D20"),("E20", "2. Tünel", "E20:F20"),
    ("C21", "R", ""),("D21", "k", ""),
    ("A24", "1. TÜNEL DEPREM YÜKLERİ", "A24:H24"),
    ("A25", "S", ""), ("B25", "C", ""), ("C25", "Y", ""), ("D25", "H", ""), ("E25", "PGA (72)", ""), ("F25", "PGA (2475)", ""), ("G25", "S1", ""), ("H25", "S2", ""),
    ("A28", "2. TÜNEL DEPREM YÜKLERİ", "A28:H28"),
    ("A29", "S", ""), ("B29", "C", ""), ("C29", "Y", ""), ("D29", "H", ""), ("E29", "PGA (72)", ""), ("F29", "PGA (2475)", ""), ("G29", "S1", ""), ("H29", "S2", "")
            ]
            yellow_cells = ["J13", "K13", "J17", "K17","G30","H30","G26","H26","D22","F22"]

        elif tünel_sayisi == 3:
            file_name = 'exact_replication_3.xlsx'
            headers = [
                    ("A1", "MALZEME TANIMI", "A1:F1"), 
    ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
    ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
    ("A3", "", ""),("B3", "", ""),("C3", "", ""),("D3", "", ""),("E3", "", ""),("F3", "", ""),
    ("A6", "1. Tünel Kemer", ""), ("B6", "", ""),  # Boş sütun
    ("A7", "1. Tünel İnvert", ""), ("B7", "", ""),  # Boş sütun
    ("A8", "2. Tünel Kemer", ""), ("B8", "", ""),  # Boş sütun
    ("A9", "2. Tünel İnvert", ""), ("B9", "", ""),  # Boş sütun
    ("A10", "3. Tünel Kemer", ""), ("B10", "", ""),  # Boş sütun
    ("A11", "3. Tünel İnvert", ""), ("B11", "", ""),  # Boş sütun
    ("A12", "", ""), ("B12", "", ""),  # Boş satır
    ("A13", "1. TÜNEL ZEMİN YÜKLERİ", "A13:K13"), 
    ("A14", "Φ'", ""), ("B14", "m", ""), ("C14", "b", ""), ("D14", "B", ""), ("E14", "f", ""), ("F14", "h", ""), ("G14", "Y", ""), ("H14", "Y'", ""), ("I14", "K0", ""), ("J14", "EV", ""), ("K14", "EH", ""),
    ("A15", "", ""), ("B15", "", ""), ("C15", "", ""), ("D15", "", ""), ("E15", "", ""), ("F15", "", ""), ("G15", "", ""), ("H15", "", ""), ("I15", "", ""), ("J15", "", ""), ("K15", "", ""),
    ("A17", "2. TÜNEL ZEMİN YÜKLERİ", "A17:K17"), 
    ("A18", "Φ'", ""), ("B18", "m", ""), ("C18", "b", ""), ("D18", "B", ""), ("E18", "f", ""), ("F18", "h", ""), ("G18", "Y", ""), ("H18", "Y'", ""), ("I18", "K0", ""), ("J18", "EV", ""), ("K18", "EH", ""),
    ("A19", "", ""), ("B19", "", ""), ("C19", "", ""), ("D19", "", ""), ("E19", "", ""), ("F19", "", ""), ("G19", "", ""), ("H19", "", ""), ("I19", "", ""), ("J19", "", ""), ("K19", "", ""),
    ("A21", "3. TÜNEL ZEMİN YÜKLERİ", "A21:K21"), 
    ("A22", "Φ'", ""), ("B22", "m", ""), ("C22", "b", ""), ("D22", "B", ""), ("E22", "f", ""), ("F22", "h", ""), ("G22", "Y", ""), ("H22", "Y'", ""), ("I22", "K0", ""), ("J22", "EV", ""), ("K22", "EH", ""),
    ("A23", "", ""), ("B23", "", ""), ("C23", "", ""), ("D23", "", ""), ("E23", "", ""), ("F23", "", ""), ("G23", "", ""), ("H23", "", ""), ("I23", "", ""), ("J23", "", ""), ("K23", "", ""),
    ("A25", "Yatak Kaysayıları", "A25:H25"),
    ("A27", "E' (kPa)", ""), ("B27", "v", ""), 
    ("C26", "1. Tünel", "C26:D26"), ("C27", "R", ""), ("D27", "k", ""), ("C28", "", ""), 
    ("E26", "2. Tünel", "E26:F26"), ("E27", "R", ""), ("F27", "k", ""), ("E28", "", ""),
    ("G26", "3. Tünel", "G26:H26"), ("G27", "R", ""), ("H27", "k", ""), ("G28", "", ""),
    ("A28", "", ""), ("B28", "", ""), ("C28", "", ""), ("D28", "", ""), ("E28", "", ""), ("F28", "", ""), ("G28", "", ""), ("H28", "", ""),
    ("A30", "1. TÜNEL DEPREM YÜKLERİ", "A30:H30"),
    ("A31", "S", ""), ("B31", "C", ""), ("C31", "Y", ""), ("D31", "H", ""), ("E31", "PGA (72)", ""), ("F31", "PGA (2475)", ""), ("G31", "S1", ""), ("H31", "S2", ""),
    ("A32", "", ""), ("B32", "", ""), ("C32", "", ""), ("D32", "", ""), ("E32", "", ""), ("F32", "", ""), ("G32", "", ""), ("H32", "", ""),
    ("A34", "2. TÜNEL DEPREM YÜKLERİ", "A34:H34"),
    ("A35", "S", ""), ("B35", "C", ""), ("C35", "Y", ""), ("D35", "H", ""), ("E35", "PGA (72)", ""), ("F35", "PGA (2475)", ""), ("G35", "S1", ""), ("H35", "S2", ""),
    ("A36", "", ""), ("B36", "", ""), ("C36", "", ""), ("D36", "", ""), ("E36", "", ""), ("F36", "", ""), ("G36", "", ""), ("H36", "", ""),
    ("A38", "3. TÜNEL DEPREM YÜKLERİ", "A38:H38"),
    ("A39", "S", ""), ("B39", "C", ""), ("C39", "Y", ""), ("D39", "H", ""), ("E39", "PGA (72)", ""), ("F39", "PGA (2475)", ""), ("G39", "S1", ""), ("H39", "S2", ""),
    ("A40", "", ""), ("B40", "", ""), ("C40", "", ""), ("D40", "", ""), ("E40", "", ""), ("F40", "", ""), ("G40", "", ""), ("H40", "", ""),
         ]

            yellow_cells = ["K15", "J19", "K23", "J15", "J23","H40","H36", "G36","K19","G40","D28","F28","H28","G32","H32"]

        elif tünel_sayisi == 4:
            file_name = 'exact_replication4.xlsx'
            headers = [
    ("A1", "MALZEME TANIMI", "A1:F1"), 
    ("A2", "YÖNETMELİK", ""), ("B2", "BETON SINIFI", ""), ("C2", "FCK", ""), ("D2", "Poisson", ""), ("E2", "TERMAL", ""), ("F2", "E", ""),
    ("A5", "KESİT KALINLIKLARI (cm)", "A5:B5"), 
    ("A6", "1. Tünel Kemer", ""), ("B6", "", ""),  # Boş sütun
    ("A7", "1. Tünel İnvert", ""), ("B7", "", ""),  # Boş sütun
    ("A8", "2. Tünel Kemer", ""), ("B8", "", ""),  # Boş sütun
    ("A9", "2. Tünel İnvert", ""), ("B9", "", ""),  # Boş sütun
    ("A10", "3. Tünel Kemer", ""), ("B10", "", ""),  # Boş sütun
    ("A11", "3. Tünel İnvert", ""), ("B11", "", ""),  # Boş sütun
    ("A12", "4. Tünel Kemer", ""), ("B12", "", ""),  # Boş sütun
    ("A13", "4. Tünel İnvert", ""), ("B13", "", ""),  # Boş sütun
    ("A15", "1. TÜNEL ZEMİN YÜKLERİ", "A15:K15"), 
    ("A16", "Φ'", ""), ("B16", "m", ""), ("C16", "b", ""), ("D16", "B", ""), ("E16", "f", ""), ("F16", "h", ""), ("G16", "Y", ""), ("H16", "Y'", ""), ("I16", "K0", ""), ("J16", "EV", ""), ("K16", "EH", ""),
    ("A17", "", ""), ("B17", "", ""), ("C17", "", ""), ("D17", "", ""), ("E17", "", ""), ("F17", "", ""), ("G17", "", ""), ("H17", "", ""), ("I17", "", ""), ("J17", "", ""), ("K17", "", ""),
    ("A19", "2. TÜNEL ZEMİN YÜKLERİ", "A19:K19"), 
    ("A20", "Φ'", ""), ("B20", "m", ""), ("C20", "b", ""), ("D20", "B", ""), ("E20", "f", ""), ("F20", "h", ""), ("G20", "Y", ""), ("H20", "Y'", ""), ("I20", "K0", ""), ("J20", "EV", ""), ("K20", "EH", ""),
    ("A21", "", ""), ("B21", "", ""), ("C21", "", ""), ("D21", "", ""), ("E21", "", ""), ("F21", "", ""), ("G21", "", ""), ("H21", "", ""), ("I21", "", ""), ("J21", "", ""), ("K21", "", ""),
    ("A23", "3. TÜNEL ZEMİN YÜKLERİ", "A23:K23"), 
    ("A24", "Φ'", ""), ("B24", "m", ""), ("C24", "b", ""), ("D24", "B", ""), ("E24", "f", ""), ("F24", "h", ""), ("G24", "Y", ""), ("H24", "Y'", ""), ("I24", "K0", ""), ("J24", "EV", ""), ("K24", "EH", ""),
    ("A25", "", ""), ("B25", "", ""), ("C25", "", ""), ("D25", "", ""), ("E25", "", ""), ("F25", "", ""), ("G25", "", ""), ("H25", "", ""), ("I25", "", ""), ("J25", "", ""), ("K25", "", ""),
    ("A27", "4. TÜNEL ZEMİN YÜKLERİ", "A27:K27"), 
    ("A28", "Φ'", ""), ("B28", "m", ""), ("C28", "b", ""), ("D28", "B", ""), ("E28", "f", ""), ("F28", "h", ""), ("G28", "Y", ""), ("H28", "Y'", ""), ("I28", "K0", ""), ("J28", "EV", ""), ("K28", "EH", ""),
    ("A29", "", ""), ("B29", "", ""), ("C29", "", ""), ("D29", "", ""), ("E29", "", ""), ("F29", "", ""), ("G29", "", ""), ("H29", "", ""), ("I29", "", ""), ("J29", "", ""), ("K29", "", ""),
    ("A31", "Yatak Kaysayıları", "A31:J31"),
    ("A33", "E' (kPa)", ""), ("B33", "v", ""), 
    ("C32", "1. Tünel", "C32:D32"), ("C33", "R", ""), ("D33", "k", ""), ("C34", "", ""), 
    ("E32", "2. Tünel", "E32:F32"), ("E33", "R", ""), ("F33", "k", ""), ("E34", "", ""),
    ("G32", "3. Tünel", "G32:H32"), ("G33", "R", ""), ("H33", "k", ""), ("G34", "", ""),
    ("I32", "4. Tünel", ""),
    ("I33", "R", ""), ("J33", "k", ""), ("I34", "", ""), ("J32", "", ""), ("J34", "", ""),
    ("A3", "", ""), ("B3", "", ""), ("C3", "", ""), ("D3", "", ""),("E3", "", ""),  ("F3", "", ""), 
    ("A36", "1. TÜNEL DEPREM YÜKLERİ", "A36:H36"),
    ("A37", "S", ""), ("B37", "C", ""), ("C37", "Y", ""), ("D37", "H", ""), ("E37", "PGA (72)", ""), ("F37", "PGA (2475)", ""), ("G37", "S1", ""), ("H37", "S2", ""),
    ("A38", "", ""), ("B38", "", ""), ("C38", "", ""), ("D38", "", ""), ("E38", "", ""), ("F38", "", ""), ("G38", "", ""), ("H38", "", ""),
    ("A40", "2. TÜNEL DEPREM YÜKLERİ", "A40:H40"),
    ("A34", "", ""),("B34", "", ""),("D34", "", ""),("F34", "", ""),("H34", "", ""),("A38", "", ""),
    ("A41", "S", ""), ("B41", "C", ""), ("C41", "Y", ""), ("D41", "H", ""), ("E41", "PGA (72)", ""), ("F41", "PGA (2475)", ""), ("G41", "S1", ""), ("H41", "S2", ""),
    ("A42", "", ""), ("B42", "", ""), ("C42", "", ""), ("D42", "", ""), ("E42", "", ""), ("F42", "", ""), ("G42", "", ""), ("H42", "", ""),
    ("A44", "3. TÜNEL DEPREM YÜKLERİ", "A44:H44"),
    ("A45", "S", ""), ("B45", "C", ""), ("C45", "Y", ""), ("D45", "H", ""), ("E45", "PGA (72)", ""), ("F45", "PGA (2475)", ""), ("G45", "S1", ""), ("H45", "S2", ""),
    ("A46", "", ""), ("B46", "", ""), ("C46", "", ""), ("D46", "", ""), ("E46", "", ""), ("F46", "", ""), ("G46", "", ""), ("H46", "", ""),
    ("A48", "4. TÜNEL DEPREM YÜKLERİ", "A48:H48"),
    ("A49", "S", ""), ("B49", "C", ""), ("C49", "Y", ""), ("D49", "H", ""), ("E49", "PGA (72)", ""), ("F49", "PGA (2475)", ""), ("G49", "S1", ""), ("H49", "S2", ""),
    ("A50", "", ""), ("B50", "", ""), ("C50", "", ""), ("D50", "", ""), ("E50", "", ""), ("F50", "", ""), ("G50", "", ""), ("H50", "", ""),
            ]

            yellow_cells = ["K16", "J20", "K24", "J16", "J24","H40", "D33","F33","H33","J33","H36", "G36", "H48", "G48","K20","G44","H44","G40","J28","K28"]

        else:
            print("Yanlış tünel girişi yapıldı")
            return  

        file_path = os.path.join(os.getcwd(), file_name)

        # Excel dosyasını yükleme veya oluşturma
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
        else:
            workbook = Workbook()

        sheet = workbook.active

        # Stil tanımları
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        subheader_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Verileri alma
        beton_sinifi = self.comboBox_beton.currentText()
        yonetmelik = self.comboBox_yonetmelik.currentText()
        tünel_verileri = [(self.tunnel_inputs[i][1].text(), self.tunnel_inputs[i + 1][1].text()) for i in range(0, tünel_sayisi * 2, 2)]

        # Sadece rakamları almak için regex kullanma
        fck_value = int(''.join(filter(str.isdigit, beton_sinifi))) if beton_sinifi else 0

        # Elastisite modülünü hesaplama
        e_value = self.get_elasticity_modulus(beton_sinifi, yonetmelik)

        # Mevcut başlıklar ve hücrelerin stil ve değerlerini uygulama
        for cell, value, merge in headers:
            sheet[cell] = value
            if "MALZEME TANIMI" in value or "KESİT KALINLIKLARI (cm)" in value or "TÜNEL ZEMİN YÜKLERİ" in value or "Yatak Kaysayıları" in value or "TÜNEL DEPREM YÜKLERİ" in value:
                sheet[cell].fill = header_fill
            else:
                sheet[cell].fill = subheader_fill
            sheet[cell].font = bold_font
            sheet[cell].alignment = center_alignment
            sheet[cell].border = thin_border
            if merge:
                sheet.merge_cells(merge)

        # FCK ve E hücrelerini doldurma
        sheet["A3"]=yonetmelik
        sheet["B3"]=beton_sinifi 
        sheet["C3"] = fck_value
        sheet["F3"] = e_value
        sheet["D3"] = 0.2  # Poisson oranı değeri
        sheet["E3"] = 0.00001  # Termal değer
        
        # Tünel verilerini yazma
        for idx, (kemer, invert) in enumerate(tünel_verileri):
            kemer_hucre = f"B{6 + idx * 2}"  # 1. Tünel Kemer hücresi
            invert_hucre = f"B{7 + idx * 2}"  # 1. Tünel İnvert hücresi

            kemer_hucre_obj = sheet[kemer_hucre]
            invert_hucre_obj = sheet[invert_hucre]

            start_kemer_hucre = self.get_start_cell(sheet, kemer_hucre_obj)
            start_invert_hucre = self.get_start_cell(sheet, invert_hucre_obj)

            start_kemer_hucre.value = kemer
            start_invert_hucre.value = invert

        # Kolon genişliklerini ayarlama
        column_widths = {'A': 20, 'B': 20, 'C': 12, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        # Sarı renkli hücreleri doldurma
        for cell in yellow_cells:
            sheet[cell].fill = yellow_fill

        # Çalışma kitabını kaydetme
        workbook.save(file_path)

        print(f"Biçimlendirilmiş Excel sayfası {file_path} konumunda oluşturuldu.")

        # Excel dosyasını açma
        if os.name == 'posix':  # macOS veya Linux
            subprocess.run(['open', file_path])
        elif os.name == 'nt':  # Windows
            os.startfile(file_path)

# Uygulamayı başlatma
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
